import numpy as np
import pandas as pd
import scipy.stats as ss
import xlrd
import openpyxl
import os
import sklearn
from datetime import timedelta, datetime
from dateutil.relativedelta import relativedelta
import time
import re
import pdb
from collections import OrderedDict, namedtuple
import pickle

os.chdir("C:\\Users\\student.DESKTOP-UT02KBN\\Desktop\\Stone_Presidio")
import defs
import cme_scrapper

#%%import Data from Com pricing spreadsheet
def int_col(n,base=26):
    "convert 0 index column number to excel column name"
    if n < 0:
        return ""
    elif n < base: 
        return chr(65 + n)         
    return int_col(n//base-1,base) + int_col(n%base,base)

def make_blb_book_row_aligned(file_name):
    """Takes file name, of (Bloomberg formatted security columns),
    Writes book to make all sheet's securities's columns date aligned.
    iterates over worksheets calling make_blb_sheet_aligned
    """
    bk_open = False
    xlrd_bk = xlrd.open_workbook(file_name)
    
    for sht_ix, name in enumerate(xlrd_bk.sheet_names()):
        b = xlrd_bk.sheet_by_index(sht_ix)       
        sz = len([i for i in b.row_values(6)])
        #See blb_dates_are_skipped()
        date_cols = [[datetime(*xlrd.xldate_as_tuple(j,0)) 
                        for j in b.col_values(i)[7:] 
                    if j] 
                        for i in range(0,sz,4)]
        dates = sorted(set([i
                            for c in date_cols
                                for i in c]),
                        reverse = True)
        is_date_aligned = [col[-1] == dates[len(col)-1]
                               for col in date_cols]
    
        if not all(is_date_aligned):
            print(f"\nWARNING: Row's not date aligned, some days missing, for {name}")
            bad_date_cols = [4*ix for ix,i in enumerate(is_date_aligned) 
                                if not i]
            print(bad_date_cols)
            if not bk_open:
                wrote_bk = openpyxl.load_workbook(file_name, data_only = True)
                bk_open = True
            #keeps book open in case have to write to another sheet; opening takes forever
            wrote_bk = make_blb_sheet_aligned(file_name, 
                                                 sht_ix, 
                                                 bad_date_cols,
                                                 dates,
                                                 wrote_bk)
    if bk_open:
        wrote_bk.save(prices_file + "(2)")
        wrote_bk.close()
    
def make_blb_sheet_aligned(file_name, sht_ix, bad_date_cols, target_indx, wrote_bk):
    """for each sheet in this file of bloomberg formatted data, all columns will
        be row aligned on each date. Must be xlsx workbook.
        file_name: name of book this operation is to be performed on
        sht_ix: sheet_ix
        bad_date_cols: list of columns of dates(as ints) which needs to be extended+next 2 cols
        target_indx: the dates that data_ix will now hold (as dt objects)
                     a strict superset of columns in bad_date_cols 
                        (values only in bad_date_cols will be removed)
        wrote_bk: the openpyxl book to write to
        returns: the book that was opened, *MUST* be {{saved + closed}} by caller
        """
    sheet = wrote_bk.worksheets[sht_ix]
    for c in bad_date_cols:
        #Have to remove all values that aren't in target_indx
        valid_idx, d_col = list(zip(*[(ix, v.value) 
                                for ix, v in enumerate(sheet[int_col(c)])
                                    if v.value in target_indx]
                                ))
        d_col = list(d_col)
        valid_idx = set(valid_idx)
        t_col = [v.value for ix, v in enumerate(sheet[int_col(c+1)])
                                    if ix in valid_idx]
        v_col = [v.value for ix, v in enumerate(sheet[int_col(c+2)])
                                    if ix in valid_idx]
        for ix, v in enumerate(target_indx):
            if ix >= len(d_col) or (d_col[ix] !=v and d_col[ix] != "#N/A N/A"): 
                try:
                    print(f"inserted {v} @ {ix}, col {c}, was {d_col[ix]}")
                except:
                    print(f"inserted {v} @ end, col {c}")
                d_col.insert(ix, v)
                t_col.insert(ix, "#N/A N/A")
                v_col.insert(ix, "#N/A N/A")
            if len(d_col) > len(target_indx):
                pdb.set_trace()
        t_col += [ "#N/A N/A"]*(len(d_col) - len(t_col))#??, not sure why not same sz
        v_col += [ "#N/A N/A"]*(len(d_col) - len(v_col))
    
        #writeback uses excel locations, 1-indexed instead of 0
        for row in range(8, 8 + len(d_col)):
            sheet.cell(column=c+1, row=row).value = d_col[row-8]
            sheet.cell(column=c+2, row=row).value = t_col[row-8]
            sheet.cell(column=c+3, row=row).value = v_col[row-8]
    return wrote_bk

def sort_contracts(sec_names):
    "Sorts Contracts, Most reccent first"
    sec_mt , sec_yr = list(zip(*[re.findall("([a-zA-Z])(\d+)", i)[0]
                                 for i in sec_names]))
    
    assert all([len(i) == 2 or i in ('1', '0', '9') for i in sec_yr]), \
            f"Recheck date conversion, {sec_yr[:100]}"
    sec_yr = [21 if yr == '1' else 
              20 if yr == '0' else 
              19 if yr == '9' else
              int(yr) 
                  for yr in sec_yr]
    month_abreviations = list(defs.con_month_abv.values())
    sec_mt = [month_abreviations.index(i) 
                  for i in sec_mt]
    reformatted = [yr*100 + m 
                       for yr,m in zip(sec_yr, sec_mt)]
    return [s for _,s in sorted(zip(reformatted, sec_names),
                                     key = lambda i: i[0],
                                     reverse= True)]

def return_expired(sec_list, curve_prices):
    """Returns a pandas df with the expiry prices for each contract, with columns
    being the number of months since that expiry. Use for predicting future prices
    without concurrent trading, eg. prediciting August settle given June settle vs
    estimating current CL1 given current CL2.
    Warning: Off by 1 Error, If contract had last trading day as last date 
    collected that contract won't be included"""
    contract_abv = re.sub("\d+", "", sec_list[0].name)[:-1].strip()
    
    collection_date = max([max(i.index) for i in sec_list])
    expired_contract = [i for i in sec_list
                            if max(i.index)<collection_date ]
    
    sorted_contract = [i.name for i in sorted(expired_contract, 
                                              key = lambda i: max(i.index),
                                              reverse = True)]
    
    prev_contract = {sorted_contract[i]: sorted_contract[i+1]
                                 for i in range(len(sorted_contract)-1)}
    prev_contract[sorted_contract[-1]] = 'last'    
    con_2_expiry = {i.name: max(i.index) for i in expired_contract}
    con_2_expiry['last'] =  min(
                                next(
                                    filter(lambda i: i.name == sorted_contract[-1], 
                                            sec_list)
                                     ).index
                                )
    #inclusive range of dates where front month
    con_range = {curr: (con_2_expiry[prev]+timedelta(1), con_2_expiry[curr]) 
                      for curr, prev in prev_contract.items()}
    
    def _expiry_prices(tick, out_sz=None):
        """"gets the expiry month prices for a given contract as np.array"""
        indxs = np.logical_and(curve_prices.index >= con_range[tick][0],
                                curve_prices.index <= con_range[tick][1])
        out = curve_prices.iloc[indxs, 0]
        if out_sz:
            missing_sz = out_sz - len(out)
            if missing_sz > 0:
                return np.concatenate((out.values, 
                                      np.repeat(np.nan, missing_sz)))
            else:
                return out[:out_sz].values
        else:
            return out.values
        
    def _make_filler(t,ix):
        "NAs to add to 'back' of df for things expired farther in the past"
        return [[np.nan]*len( _expiry_prices(t)) for _ in range(ix)]

    expired_curve = pd.DataFrame(
                        np.concatenate(
                            [np.stack(
                                [_expiry_prices(j,
                                           out_sz = len( _expiry_prices(t)))
                                     for j in sorted_contract[ix:]]
                                + _make_filler(t,ix)
                                 ).T
                                    for ix, t in enumerate(sorted_contract)
                            ]),
                        index = [f"{t} {i} before expiry"
                                     for t in sorted_contract
                                         for i in range(len(_expiry_prices(t)))],
                        columns = [f"{contract_abv} {i}Ago" #not all contracts 1Mo
                                   for i in range(len(sorted_contract))]
                        )
    return expired_curve

#%%
def get_blb_excel(prices_file, individual_securities = True,  already_formatted = False):
    """Reads in data from an excel workbook with many sheets of Bloomberg formated
    historical future's prices, with Top Row Dates aligned.
    Returns Pandas DF's
    """
    if not already_formatted:
        print(f"Will be long, is reformatting, make sure {prices_file} is closed")
        make_blb_book_row_aligned(prices_file)#modifies book

    xl_bk = xlrd.open_workbook(prices_file)
    commodities = xl_bk.sheet_names()
    security_l = []
    curve_prices_l = []
    expired_curves_d = {}
    
    for sht_ix, name in enumerate(commodities):
        b = xl_bk.sheet_by_index(sht_ix)               
        com_maturities = [i for i in b.row_values(0)[1::4] if i]
        sz = len([i for i in b.row_values(6)])
        date_cols = [[datetime(*xlrd.xldate_as_tuple(j,0)) 
                            for j in b.col_values(i)[7:] 
                        if j] 
                            for i in range(0,sz,4)]
        dates = sorted(set([i
                            for c in date_cols
                                for i in c]),
                        reverse = True)
        curve_prices = pd.DataFrame(np.array([pd.to_numeric(b.col_values(i)[7:],
                                                            errors = 'coerce'
                                                            )
                                                for i in range(2,sz,4)
                                                ]).T,
                                     columns = com_maturities,
                                     index = dates,
                                     )            
        curve_prices_l += [curve_prices]#.dropna()

        ##individual_securities
        sec_df = pd.DataFrame(list(zip(*[b.col_values(i)[7:]
                                                for i in range(1,sz,4)]
                                                )) )
        securities = np.unique(sec_df.values)
        securities = securities[~np.isin(securities, ('', '#N/A N/A'))]
                                                      
        def get_security(s):
            "returns prices for 1 single future"
            rows, _ = np.where(sec_df == s)
            #get all of row, column not just those indexes.
            df = pd.Series(data = curve_prices.values[sec_df == s], 
                      index = curve_prices.index[rows],
                      name = s
                      ).dropna()
            return df[~pd.isna(df.index)]
        sec_list = list(filter(lambda i: len(i) > 0,
                               [get_security(s) for s in securities]))
        security_l += sec_list
    
        ##expired securities
        contract_abv = re.sub("\d+", "", sec_list[0].name)[:-1].strip()
        expired_curve = return_expired(sec_list,
                                       curve_prices)
        expired_curves_d[contract_abv] = expired_curve

    #make dataframes
    curve_prices_df = pd.concat(curve_prices_l,
                                axis=1,
                                join='outer', 
                                sort=True).iloc[::-1]#largest axis up top
    curve_prices_df.dropna(how='all', inplace=True)
    curve_prices_df.columns = [i.replace("COMB", "").replace("Comdty", "").replace(" ", "") 
                for i in curve_prices_df.columns]#eg CL 1
    securities_df = pd.concat(security_l, 
                              axis=1, 
                              join='outer',
                              sort=True).iloc[::-1]
    #sizes are off since Some columns have date's with prices, but no ticker??
    securities_df.dropna(how='all', inplace=True)
    # expired_curves_df = pd.concat(expired_curve_d, 
    #                               axis=1, 
    #                               join='outer').iloc[::-1]
    return curve_prices_df, securities_df, expired_curves_d

def save_struct(struct, name):
    "handler to pickle data"    
    os.chdir("C:\\Users\\student.DESKTOP-UT02KBN\\Desktop\\Stone_Presidio\\Data\\pickled_data")
    with open(f'{name}.p', 'wb') as file:
        pickle.dump(struct,  file)
    
def load_struct(name):
    os.chdir("C:\\Users\\student.DESKTOP-UT02KBN\\Desktop\\Stone_Presidio\\Data\\pickled_data")
    with open(f'{name}.p', 'rb') as file:
        return pickle.load(file)
    
def data_handler(save_data = False):
    """"MODIFIES GLOBALS; by assigning values to data_structs
    All purpose data handler, uses subfolder 'picked_data'.
    """
    data_structs = ('curve_prices_df',
                    'securities_df', 
                    'expired_curves_d', 
                    'cme_df',
                    'eia_bio_df')
    if save_data:
        for struct_name in data_structs:
            try:
                struct = eval(struct_name)
                save_struct(struct, struct_name)
            except Exception as e:
                print(e)
        return None
    else:#load data
        for struct_name in data_structs:
           try:
               exec(f"global {struct_name}", globals())
               exec(f"globals()[struct_name] = load_struct(struct_name)")
           except Exception as e:
               print(f"""Haven't pickeled {struct_name}, Processing data Now.\n""", e, "\n\n")
               
               if struct_name in ('curve_prices_df', 'securities_df', 'expired_curves_d'):
                   os.chdir("C:\\Users\\student.DESKTOP-UT02KBN\\Desktop\\Stone_Presidio\\Data")
                   prices_file = "16.16 Historical Commodity Price Data.xlsx"
                   (curve_prices_df, 
                    securities_df, 
                    expired_curves_d) = get_blb_excel(prices_file, 
                                                      already_formatted  = True)                                                     
                   save_struct(curve_prices_df, 'curve_prices_df')
                   save_struct(securities_df, 'securities_df')
                   save_struct(expired_curves_d, 'expired_curves_d')
                   
               elif struct_name == 'cme_df':
                   cme_df = cme_scrapper.make_cme_df(cme_scrapper.get_all_cme_prx())
                   save_struct(cme_df,'cme_df')
                   
               elif struct_name == 'eia_bio_df':
                   os.chdir("C:\\Users\\student.DESKTOP-UT02KBN\\Desktop\\Stone_Presidio\\Data")
                   eia_bio_df = cme_scrapper.eia_renewable_table(table=17)
                   save_struct(eia_bio_df, 'eia_bio_df')
                   
               exec(f"global {struct_name}", globals())
               exec(f"globals()[struct_name] = load_struct(struct_name)") 
    
data_handler(save_data = False)
#%%    
# futures = [i.replace("COMB", "").replace("Comdty", "").replace(" ", "") 
#             for i in curve_prices_df]#eg CL 1
# futures_ab = set([re.sub("\d+", "",i) 
#                     for i in futures])#eg CL

# month_rng = sorted(set([datetime(d.year, d.month, 1)
#                         for d in securities_df.index]))
# month_abvs = sort_contracts([f"{defs.int_month_abv[i.month]}{str(i.year)[-2:]}" 
#                              for i in month_rng])
#%% 
=

#%%






#%% Predicting comodity pries
#impute missing
curve_prices_df = curve_prices_df.dropna(axis=0)
a=(curve_prices_df =='#N/A N/A')
curve_prices_df = curve_prices_df[~a.any(axis=1)]

#%%
front_cols = [i for i in curve_prices_df.columns 
             if '1 ' in i and '11' not in i and '21' not in i]
rest_cols = [i for i in curve_prices_df.columns if i not in front_cols]

X = curve_prices_df.loc[:,rest_cols]
Y = curve_prices_df.loc[:,front_cols[0]]
#%%
from sklearn.pipeline import Pipeline
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LinearRegression
from sklearn.impute import SimpleImputer
from sklearn.preprocessing import StandardScaler

X_train, X_test, y_train, y_test = train_test_split(X, Y, random_state=0)
pipe = Pipeline([('impute', SimpleImputer()),
                     ('scaler', StandardScaler()),
                     ('linear', LinearRegression())])

pipe.fit(X_train, y_train)

#%%
#curve_prices_df = pd.concat(curve_prices_d, axis=1)

#%%
#Predicting revenue for each of the companies, by reading in revenue from model

#import pyxlsb
program_Cos = ['Express Grain (KCO)']
fatoil_Cos = ['Western Dubuque', 'Hero', 'Kolmar', 'Sinclair', 'Mendota', 'Verbio']
energy_Cos = ['SGR Energy', 'Hiper Gas', 'Petromax']
cocoa_Cos = ['Hershey']
Cos =  [program_Cos, fatoil_Cos, energy_Cos, cocoa_Cos]

sheet_names = ['Rev Build_ Programs', 'Rev Build_ Fats&Oils', 
               'Rev Build_ Energy', 'Rev Build_ Cocoa Trading']
hist_ixs = [slice(12, 21), slice(12, 21), slice(12, 21), slice(12, 21)]

model_file_dir = 'C:\\Users\\student.DESKTOP-UT02KBN\\Desktop\\Stone_Presidio'
os.chdir(model_file_dir)
model_file_name = '20200520 02 FCStone Presidio Model WIP.xlsx'
model_bk = xlrd.open_workbook(model_file + "\\" + model_file_name)
#print(model_bk.sheet_names())    
#model_pybk = pyxlsb.open_workbook(model_file + "\\" + model_file_name)
#print(model_pybk.sheets)
for sheet_name, h_ix, sector_Cos in zip(sheet_names, hist_ixs, Cos):
    b = model_bk.sheet_by_name(sheet_name)
    co = sector_Cos[0]
    co_ix = b.col_values(0).index(co)    
    historical_rev = b.row_values(co_ix)[h_ix]
    
    try:
        gross_ix =  b.col_values(0, 
                                start_rowx = 0, 
                                end_rowx = 200).index('Gross Revenue')
        cos_ix =  b.col_values(0, 
                                start_rowx = 0, 
                                end_rowx = 200).index('Cost of Sales')
        margin = [1-j/i for i,j in zip(b.row_values(gross_ix)[h_ix],
                                     b.row_values(cos_ix)[h_ix])]#COS is line below gross revenue
    except:
        print(b.col_values(0, start_rowx = 0, end_rowx = 200), "\n\n\n\n")
        margin = None
    print(margin)
#for co in fatoil_Cos: 
#%%
    


#%% Make Price Distribution Graphs from future's
from scipy import integrate
import defs

hist_val = curve_prices_df.reindex(curve_prices_df.index[::-1])
hist_pct = hist_val.pct_change(periods = 252)#6mo out
hist_vol = hist_pct.std(axis=0)

for ab in defs.abv_cme_url:
    if ab:
        tick = ab + "1"
        ix = futures.index(tick)
        base_ix = hist_val.iloc[:,ix].first_valid_index()
        base_prx = hist_val.loc[base_ix, tick]
        dist = ss.lognorm(loc = 0, scale = 1, s = hist_vol.values[ix])
        
        x = np.linspace(hist_pct.iloc[:,ix].min(),
                        hist_pct.iloc[:,ix].max(),
                        100)*2
        
        fig, ax =plt.subplots()
        x_axis = [i for i in  base_prx * (1 + x) if i > 0]
        
        probs = [integrate.quad(lambda x: dist.pdf(x/base_prx)/base_prx, i,j)[0]
                         for i,j in zip(x_axis[:-1], x_axis[1:])]
        probs = [100*i/sum(probs) for i in probs]
        
        plt.bar(x_axis[:-1], probs, label="Probabilities")
        plt.legend()
        plt.title(f"Probability of ending prices of {hist_vol.index[ix]} in 12mo")
        xformatter = mticker.FormatStrFormatter('$%.0f')
        ax.xaxis.set_major_formatter(xformatter)
        yformatter = mticker.FormatStrFormatter('%.1f%%')
        ax.yaxis.set_major_formatter(yformatter)
        plt.show()


#plt.plot(x_axis, dist.pdf(x), label='dist')
#plt.yscale('log')