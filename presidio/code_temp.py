import numpy as np
import scipy.stats as ss
import matplotlib.pyplot as plt
import pandas as pd

data = pd.read_csv("file:///C:/Users/student.DESKTOP-UT02KBN/Downloads/Ho5A5e.csv")
(nrow,ncol) = data.shape
x = np.array(pd.concat([pd.DataFrame([1]*nrow), data.iloc[:,:3]],axis=1))
y = np.array(data.loc[:,'y'])
theta = np.ones((4,1))

def prediction_reg(x,theta):
    pred = x @ theta
    return pred.reshape(-1)



prediction_reg(x,theta)[:2]
#%%
def calc_loss(theta, data):
    (nrow,ncol) = data.shape
    x = np.array(pd.concat([pd.DataFrame([1]*nrow), data.iloc[:,:3]],axis=1))
    y = np.array(data.loc[:,'y'])
    pred = prediction_reg(x, theta)
    loss = np.mean((y-pred)**2)
    return loss

calc_loss(theta, data[:2])

#%%
def calc_grad(train, theta):
    (nrow,ncol) = data.shape
    x = np.array(pd.concat([pd.DataFrame([1]*nrow), data.iloc[:,:3]],axis=1))
    y = np.array(data.loc[:,'y'])
    theta = np.ones((4,1))
#    print(x.shape, theta.shape, theta)
    pred = prediction_reg(x, theta)
    g1 = 2*(y-pred).reshape(1,-1)
    grad = theta @ g1
    return np.mean(grad, axis = 1)

calc_grad(theta, data[:2])

#%%

data = pd.read_csv("file:///C:/Users/student.DESKTOP-UT02KBN/Downloads/VAIwT1.csv")
data = data[list(data.loc[:,'TV']<10**10)]

#fig, ax =plt.subplot(1,1)
for name in data.columns[1:4]:
    plt.scatter(data.loc[:,name], data.loc[:,'sales'])
    plt.xlabel(name)
    plt.ylabel('Sales')
    plt.title(f"Corr: {data.loc[:,name].corr(data.loc[:,'sales'])}")
    plt.show()

data['CPS'] = data.apply(lambda row: sum(row[1:4])/row[4],axis=1)
data['isSuccess'] = data.apply(lambda row: (row[4]>15) and (row[5] < 20), axis=1)

data['isSuccess'].sum()
                   #%%
import sklearn as sk
from sklearn.svm import SVC
train = data[list(data['id']<=160)]
y = train['isSuccess']
x = train.iloc[:,1:4]
model = SVC(gamma = 0.1)
sk.gridsearchcv(model, {gamma:})
model.fit(x,y)

np.mean(model.predict(x) == y)
#%%
test = data[list(data['id']>160)]
y = test['isSuccess']
x = test.iloc[:,1:4]
np.mean(model.predict(x) == y)
#%%
url="https://slatestarscratchpad.tumblr.com/archive/"

from bs4 import BeautifulSoup
from urllib.request import urlopen
import re
from datetime import datetime
f_M = 7
f_Y = 2014
l_M = min(datetime.now().month + 1, 12)
l_Y = datetime.now().year

urls = []
for y,m in ((y,m) for y in range(f_Y, l_Y+1)
 for m in range((y==f_Y)*first_M or 1, ((y==l_Y)*l_M or 12) + 1)):
    print(f"trying: {url + str(y) + '/' + str(m)}")
    html_page = urlopen(url + str(y) + '/' + str(m))
    soup = BeautifulSoup(html_page)
    for link in soup.findAll('a', attrs={'href': re.compile("^http://")}):
        print(link.get('href'))
        urls.append(link.get('href'))
#%%









#%%
def int_col(n,base=26):
    "convert 0 index column number to excel column name"
    if n < 0:
        return ""
    elif n < base:
        return chr(65 + n)
    return int_col(n//base-1,base) + int_col(n%base,base)

def blb_dates_are_skipped(b):
    """Takes an xlrd worksheet (Bloomberg formated) and returns wheather the dates
    for each security column do not skip any date that is included by another
    column on that worksheet.
    b: xlrd worksheet
    Returns: (list of bools of if security is missing pricing data for a date included by another,
              all the dates every security should have)
    """
    sz = len(b.row_values(6))
    date_cols = [[datetime(*xlrd.xldate_as_tuple(j,0))
                    for j in b.col_values(i)[7:]
                if j]
                    for i in range(0,sz,4)]
    dates = sorted(set([i
                        for c in date_cols
                            for i in c]),
                    reverse = True)
    is_date_aligned = [col[-1] == dates[len(col)-1]
                           for col in date_cols]
    return is_date_aligned, dates

os.chdir("C:\\Users\\student.DESKTOP-UT02KBN\\Desktop\\Stone_Presidio\\Data")
prices_file = "16.16 Historical Commodity Price Data.xlsx"

individual_securities = True

wrote_bk = None
xl_bk = xlrd.open_workbook(prices_file)
commodities = xl_bk.sheet_names()
security_d = []
curve_prices_d = []
#for ix, name in enumerate(commodities):

sht_ix = 0
name = 'Corn'

b = xl_bk.sheet_by_index(sht_ix)

is_date_aligned, dates = blb_dates_are_skipped(b)
#%%
if not all(is_date_aligned):
    print(f"\nWARNING: Row's not date aligned, some days missing, for {name}")
    bad_date_cols = [4*ix for ix,i in enumerate(is_date_aligned)
                        if not i]
#    #keeps book open in case have to write to another sheet; opening takes forever
#    wrote_bk = make_blb_row_aligned(prices_file,
#                         sht_ix,
#                         bad_date_cols,
#                         dates,
#                         bk = wrote_bk)
#%
    (file_name, sheet_ix, bad_date_cols, target_indx, bk) =  (prices_file,
                                                                     sht_ix,
                                                                     bad_date_cols,
                                                                     dates,
                                                                     wrote_bk)
    bk = bk or openpyxl.load_workbook(file_name, data_only = True)
    sheet = bk.worksheets[sheet_ix]
    for c in bad_date_cols:
        #Have to remove all values that aren't in target_indx
        valid_idx, d_col = list(zip(*[(ix, v.value)
                                for ix, v in enumerate(sheet[int_col(c)])
                                    if v.value in target_indx]
                                ))#[7:]
        d_col = list(d_col)
        t_col = [v.value for ix, v in enumerate(sheet[int_col(c+1)])
                                    if ix in valid_idx]
        v_col = [v.value for ix, v in enumerate(sheet[int_col(c+2)])
                                    if ix in valid_idx]
        print(f"Bad col len: {len(d_col)} vs. {len(target_indx)}")
#        for i,v in enumerate(reversed(d_col)):#openpyxl appends a bunch of empty cells
#            if v is not None:
#                valid_till = len(d_col) - i
#                break
#        d_col = d_col[:valid_till]
#        t_col =[i.value for i in sheet[int_col(c+1)]][7:valid_till]
#        v_col = [i.value for i in sheet[int_col(c+2)]][7:valid_till]

        for ix, v in enumerate(target_indx):
            if d_col[ix] !=v and d_col[ix] != "#N/A N/A":
                print(f"inserted {v} @ {ix}, col {c}, was {d_col[ix]}")
                d_col.insert(ix, v)
                t_col.insert(ix, "#N/A N/A")
                v_col.insert(ix, "#N/A N/A")
            if len(d_col) > len(target_indx):
                pdb.set_trace()
#                time.sleep(0.1)
        t_col += [ "#N/A N/A"]*(len(d_col) - len(t_col))#??, not sure why not same sz
        v_col += [ "#N/A N/A"]*(len(d_col) - len(v_col))

        #writeback uses excel locations, 1-indexed not 0
        for row in range(8, 8 + len(d_col)):
            sheet.cell(column=c+1, row=row).value = d_col[row-8]
            sheet.cell(column=c+2, row=row).value = t_col[row-8]
            sheet.cell(column=c+3, row=row).value = v_col[row-8]

    wrote_bk.save(prices_file)


#%%











#%%
import numpy as np
import pandas as pd
import scipy.stats as ss
import xlrd
import openpyxl
import os
import sklearn
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import time
import re
import pdb

os.chdir("C:\\Users\\student.DESKTOP-UT02KBN\\Desktop\\Stone_Presidio\\Data")
file_name = "16.16 Historical Commodity Price Data.xlsx"

def int_col(n,base=26):
    "convert 0 index column number to excel column name"
    if n < 0:
        return ""
    elif n < base:
        return chr(65 + n)
    return int_col(n//base-1,base) + int_col(n%base,base)


#make_blb_book_row_aligned(prices_file)
def process_blb_xlsx(file_name):
    bk_open = False
    xlrd_bk = xlrd.open_workbook(file_name)

    for sht_ix, name in enumerate(xlrd_bk.sheet_names()):
        b = xlrd_bk.sheet_by_index(sht_ix)
        sz = len([i for i in b.row_values(6)])
        #See blb_dates_are_skipped()
        date_cols = [[datetime(*xlrd.xldate_as_tuple(j,0))
                        for j in b.col_values(i)[7:]
                    if j]
                        for i in range(0,sz,4)]
        dates = sorted(set([i
                            for c in date_cols
                                for i in c]),
                        reverse = True)
        is_date_aligned = [col[-1] == dates[len(col)-1]
                               for col in date_cols]

        if not all(is_date_aligned):
            print(f"\nWARNING: Row's not date aligned, some days missing, for {name}")
            bad_date_cols = [4*ix for ix,i in enumerate(is_date_aligned)
                                if not i]
            print(bad_date_cols)
            if not bk_open:
                wrote_bk = openpyxl.load_workbook(file_name, data_only = True)
                bk_open = True
            target_indx = dates
            #%make_blb_sheet_aligned
            sheet = wrote_bk.worksheets[sht_ix]
            for c in bad_date_cols:
                #Have to remove all values that aren't in target_indx
                valid_idx, d_col = list(zip(*[(ix, v.value)
                                        for ix, v in enumerate(sheet[int_col(c)])
                                            if v.value in target_indx]
                                        ))
                d_col = list(d_col)
                valid_idx = set(valid_idx)
                t_col = [v.value for ix, v in enumerate(sheet[int_col(c+1)])
                                            if ix in valid_idx]
                v_col = [v.value for ix, v in enumerate(sheet[int_col(c+2)])
                                            if ix in valid_idx]
                for ix, v in enumerate(target_indx):
                    if ix >= len(d_col) or (d_col[ix] !=v and d_col[ix] != "#N/A N/A"):
                        try:
                            print(f"inserted {v} @ {ix}, col {c}, was {d_col[ix]}")
                        except:
                            print(f"inserted {v} @ end, col {c}")
                        d_col.insert(ix, v)
                        t_col.insert(ix, "#N/A N/A")
                        v_col.insert(ix, "#N/A N/A")
                    if len(d_col) > len(target_indx):
                        pdb.set_trace()
                t_col += [ "#N/A N/A"]*(len(d_col) - len(t_col))#??, not sure why not same sz
                v_col += [ "#N/A N/A"]*(len(d_col) - len(v_col))

                #writeback uses excel locations, 1-indexed instead of 0
                for row in range(8, 8 + len(d_col)):
                    sheet.cell(column=c+1, row=row).value = d_col[row-8]
                    sheet.cell(column=c+2, row=row).value = t_col[row-8]
                    sheet.cell(column=c+3, row=row).value = v_col[row-8]
            #%end make_blb_sheet_aligned
    if bk_open:
        wrote_bk.save(file_name)
        print("Saved")
        wrote_bk.close()

#%% Price pull request
all_hist = process_macroTrendsnet()

all_hist = all_hist[['coffee', 'cocoa', 'cotton', 'corn', 'wheat',
                     'wti', 'heating']]
cotton = all_hist.pop('cotton')
cotton.columns = ['CT1']
cotton.name = 'CT1'
all_hist.columns = ["KC1", "CC1", "C1", "W1", "CL1", "HO1"]

temp = curve_prices_df.loc[:,["KC1", "CC1", "C1", "W1", "CL1", "HO1"]]

last_dt = max(temp.index)
temp = pd.concat((all_hist[all_hist.index > last_dt], temp))
temp = temp.join(cotton, how="inner")
temp = temp[temp.isna().sum(axis=1) < 3]
temp = temp.loc[:,['KC1', 'CC1', 'CT1', 'C1', 'W1', 'CL1', 'HO1']]

names_row = ["Coffee", "Cocoa (NY)", "Cotton", "Corn","Wheat (Chicago)",
                "Crude Oil", "Heating Oil"]
units_row = ['U.S. dollars and cents per pound',
             'U.S. dollars and cents per Metric Ton',
             "U.S. dollars per pound",
             'U.S. cents per bushel',
             'U.S. cents per bushel',
             'U.S. dollars and cents per barrel',
             'U.S. dollars and cents per gallon']
df = pd.DataFrame([names_row, units_row], columns = temp.columns)
temp = pd.concat((df, temp))
os.chdir("C:\\Users\\student.DESKTOP-UT02KBN\\Desktop\\Stone_Presidio\\Data")
temp.to_csv("price request 6_19.csv")

#%%

bio = eia_bio_df.loc[:,"Retail Biodiesel"]
bio.name = "Retail Biodiesel"
retail_diesel2.columns = ["No 2 Diesel Retail"]

bean_oil = process_macroTrendsnet()['soybean oil'].dropna()
bean_oil.name = "Soybean Oil"
bean_oil = bean_oil.loc[[i
                         for i in bean_oil.index
                         if i >= min(min(retail_diesel2.index),
                                     min(bio.index))]]

bio.index = [next(j
                    for j in reversed(bean_oil.index)
                    if j >= i)
             for i in bio.index]

retail_diesel2.index = [next(j
                        for j in reversed(bean_oil.index)
                        if j >= i)
                    for i in retail_diesel2.index]

temp = bean_oil.to_frame().join((bio, retail_diesel2),
                           how='left',
                           sort = True
                           ).fillna(method='bfill').fillna(method='ffill')

units_row = ["(Dollars per Pound)",
             "(Dollars per Gallon)",
             "(Dollars per Gallon)"]
# os.chdir("C:\\Users\\student.DESKTOP-UT02KBN\\Desktop\\Stone_Presidio\\Data")
# temp.to_csv("price request 6_22.csv")
#%%
def where_BLB_MacroTrends_differ():
    """There's some differences between say day pricing on
    Bloomberg and Macrotrends and I can't tell why"""
    for df_n, m_n in zip(('CC1', 'KC1', 'C1', 'HO1', 'BO1', 'W1', 'CL1'),
                        ['cocoa', 'coffee', 'corn', 'heating', 'soybean oil', 'wheat',
           'wti']):
        x = curve_prices_df[df_n].dropna()
        y = historic_front_month.loc[historic_front_month.index.isin(x.index),
                                     m_n].dropna()
        x = x[y.index]
        a = pd.DataFrame(({'x':x,
                           'y':y,
                           'year':[i.year for i in x.index],
                           'day': [i.day for i in x.index]}))
        difs =  a.apply(lambda row: row.iloc[0] != row.iloc[1],
                       axis =1)
        print(df_n,
              m_n,
                # difs.groupby(a['year']).sum(),
                difs.groupby(a['day']).sum(),
               # set([i.day for i in difs.index]),
              # np.corrcoef(x,y),
              "\n\n"
                )
    # a[a.apply(lambda row: row.iloc[0] != row.iloc[1] and row.iloc[2] == 2015,
    #           axis =1)]
# where_BLB_MacroTrends_differ()

# plt.plot(retail_diesel2)
# plt.plot(eia_bio_df.index, eia_bio_df["Retail Diesel"])


# plt.plot(retail_diesel2.join(eia_bio_df.loc[:,"Retail Diesel"],
#                             how='left',
#                             sort = True
#                             ).fillna(method='bfill')['Retail Diesel'])
# plt.plot(eia_bio_df["Retail Diesel"])
# plt.show()
epa_diesel = eia_bio_df['Retail Diesel']

epa_diesel.index = [next(j
                        for j in reversed(retail_diesel2.index)
                        if j >= i)
                    for i in epa_diesel.index]

temp = retail_diesel2.join(epa_diesel,
                           how='left',
                           sort = True
                           ).fillna(method='bfill').fillna(method='ffill')
plt.plot(temp)
plt.show()

#%%



num_cols = con_back or len(sorted_contract)
    def _expiry_prices(tick, out_sz=None):
        """"gets the expiry month prices for a given contract as np.array,
        tick is ticker of the expired contract.
        For contracts that have fewer trading days then out_sz,
        the prices in the *first* front month trading date will be used.
        This will be flipped for the very last, unexpired contract, which will
        map to first trading days of each contract(??this doesn't work??, grib)
        out_sz: size of df to be returned"""
        indxs = np.logical_and(curve_prices.index >= con_range[tick][0],
                                curve_prices.index <= con_range[tick][1])
        out = curve_prices.iloc[indxs, 0]#front month prices for this Commodity
        if out_sz:
            missing_sz = out_sz - len(out)
            if missing_sz > 0:
                #contract has fewer trading days then some contract after it
                return np.concatenate((out.values,
                                      np.repeat(out.iloc[-1], missing_sz)))
            else:
                if tick != front_mo_con:
                    return out[:out_sz].values
                else:
                    return out[-out_sz:].values
        else:
            return out.values

    def _make_filler(t,ix):
        """NAs to add to 'back' of df for contracts that weren't traded over the entire period.
        eg. to the back of CLZ18 for the dates beyond (Dec 18 - start of data)
        when this wasn't traded for the entire length of backtest
        Needs to be transposed"""
        return [[np.nan]*len( _expiry_prices(t)) for _ in range(ix - num_cols)]

    def _make_index():
        "Makes the index for ticker t of datetime objects"
        #each list(filter) will be same length as out_sz
        dt_index =  [ix for tick in sorted_contract
                         for ix in filter(lambda i: i >= con_range[tick][0]
                                                and i <= con_range[tick][1],
                                          curve_prices.index)]
        named_index = [f"{t} {i} before expiry"
                            for t in sorted_contract
                                for i in range(len(_expiry_prices(t)))]
        tuples = list(zip(dt_index, named_index))
        multi_indx = pd.MultiIndex.from_tuples(tuples,
                                               names=['Dates', 'Description'])

        axis_unique = len(np.unique(dt_index)) == len(dt_index)
        assert axis_unique, "datetime axis is not unqiue, repeated dates"
        return dt_index

    expired_curve = pd.DataFrame(
                        np.concatenate(
                            [np.stack(
                                [_expiry_prices(j,
                                                out_sz = len( _expiry_prices(t)))
                                     for j in sorted_contract[ix:ix+num_cols]]
                                + _make_filler(t,ix)
                                 ).T
                                    for ix, t in enumerate(sorted_contract):
                                        if ix < num_cols
                            ]),
                        index = _make_index(),
                        columns = [f"{contract_abv} {i}Ago" #not all contracts 1Mo
                                   for i in range(num_cols)]
                        )